"""
Test BOM Excel Üretici
20.000 satır, hiyerarşik yapı, BOM Manager uyumlu format.

Sütunlar: Level | Parça Adı | Parça No | Seri No | Adet | Jira Key | Ağırlık(kg) | Malzeme | Tedarikçi | Revizyon | Durum | Kritiklik | MTBF(saat) | Birim Fiyat | Para Birimi | Son Revizyon | Onaylayan | Kategori | Açıklama | Notlar
"""

import random
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)

# ─── Veri havuzları ───────────────────────────────────────────────
SUBSYSTEMS = [
    "Güç Sistemi", "Termal Yönetim", "Yapısal Gövde", "RF Anten Bloğu",
    "Sinyal İşleme Ünitesi", "Veri Kayıt Birimi", "Haberleşme Modülü",
    "Navigasyon Birimi", "Kontrol Elektroniği", "Yazılım ve Ara Katman",
    "Test ve Kalibrasyon", "Mekanik Aktüatörler", "Optik Sistem",
    "Soğutma ve Klima", "Güvenlik Sistemi", "Kullanıcı Arayüzü",
    "Enerji Depolama", "Kablo Demeti", "Sensör Dizisi", "Bakım Erişim Paneli",
]

ASSEMBLY_PREFIXES = [
    "Ana", "Yedek", "Birincil", "İkincil", "Harici", "Dahili",
    "Aktif", "Pasif", "Dijital", "Analog", "RF", "DC", "AC",
    "Yüksek Güçlü", "Düşük Profilli", "Modüler", "Sabit",
]

PART_TYPES = [
    "PCB Kartı", "Konnektör", "Kablo", "Rezistör", "Kondansatör",
    "Transistör", "Entegre", "Trafo", "Röle", "Sigorta",
    "Vida", "Somun", "Yay", "Conta", "Bağlantı Parçası",
    "Soğutucu", "Filtre", "İnduktör", "Diyot", "LED",
    "Switch", "Potansiyometre", "Kristal", "Osilatör", "Faz Kilitli Döngü",
    "ADC", "DAC", "FPGA", "DSP", "Mikrodenetleyici",
    "Bellek Modülü", "Flash Bellek", "EEPROM", "Optokuplör", "Varistör",
]

MATERIALS = ["Al6061", "Al7075", "Ti6Al4V", "SS304", "SS316", "Karbon Fiber", "FR4", "Rogers4003", "Kovar", "Beriliyum Bakır"]
SUPPLIERS = ["Tedarikçi-A", "Tedarikçi-B", "Tedarikçi-C", "Tedarikçi-D", "Tedarikçi-E", "Tedarikçi-F", "Tedarikçi-G"]
STATUSES = ["Onaylı", "Taslak", "İncelemede", "Askıya Alındı"]
CRITICALITIES = ["Kritik", "Yüksek", "Orta", "Düşük"]
CATEGORIES = ["Elektronik", "Mekanik", "Yazılım", "Kablo", "Standart Parça", "Özel İmalat"]
CURRENCIES = ["TL", "USD", "EUR"]

def rand_pn(prefix, index):
    return f"{prefix}-{index:05d}-{random.choice('ABCDE')}"

def rand_sn(index):
    return f"SN{index:07d}"

def rand_jira():
    proj = random.choice(["BOM", "HW", "SW", "ME", "SYS", "TEST"])
    return f"{proj}-{random.randint(100,9999)}"

def rand_date():
    y = random.randint(2022, 2025)
    m = random.randint(1, 12)
    d = random.randint(1, 28)
    return f"{y}-{m:02d}-{d:02d}"

def rand_approver():
    return random.choice(["A.Yılmaz", "B.Kaya", "C.Demir", "D.Şahin", "E.Çelik", "F.Arslan"])

# ─── Hiyerarşi Oluştur ────────────────────────────────────────────
# Hedef: ~20.000 satır
# Level 0: 20 alt sistem
# Level 1: her birinde 10 assembly → 200
# Level 2: her birinde 10 sub-assembly → 2.000
# Level 3: her birinde ~9 parça → ~18.000
# Toplam: 20 + 200 + 2000 + 18000 = 20.220 satır

rows = []
idx = 1

for si, subsys_name in enumerate(SUBSYSTEMS):
    # Level 0
    pn0 = rand_pn("SYS", si + 1)
    rows.append({
        "level": 0,
        "name": subsys_name,
        "pn": pn0,
        "sn": rand_sn(idx),
        "qty": 1,
        "jira": rand_jira(),
        "weight": round(random.uniform(0.5, 50.0), 3),
        "material": random.choice(MATERIALS),
        "supplier": random.choice(SUPPLIERS),
        "rev": f"Rev {random.randint(1,10):02d}",
        "status": random.choice(STATUSES),
        "criticality": random.choice(CRITICALITIES),
        "mtbf": random.randint(500, 100000),
        "unit_price": round(random.uniform(100, 50000), 2),
        "currency": random.choice(CURRENCIES),
        "last_rev": rand_date(),
        "approver": rand_approver(),
        "category": random.choice(CATEGORIES),
        "desc": f"{subsys_name} ana montaj grubu",
        "notes": f"Bkz. ICD-{random.randint(100,999)}",
    })
    idx += 1

    for ai in range(10):
        prefix_a = random.choice(ASSEMBLY_PREFIXES)
        asm_name = f"{prefix_a} {subsys_name} Asemblisi {ai+1}"
        pn1 = rand_pn("ASM", idx)
        rows.append({
            "level": 1,
            "name": asm_name,
            "pn": pn1,
            "sn": rand_sn(idx),
            "qty": random.choice([1, 1, 1, 2, 4]),
            "jira": rand_jira(),
            "weight": round(random.uniform(0.1, 10.0), 3),
            "material": random.choice(MATERIALS),
            "supplier": random.choice(SUPPLIERS),
            "rev": f"Rev {random.randint(1,5):02d}",
            "status": random.choice(STATUSES),
            "criticality": random.choice(CRITICALITIES),
            "mtbf": random.randint(1000, 50000),
            "unit_price": round(random.uniform(50, 10000), 2),
            "currency": random.choice(CURRENCIES),
            "last_rev": rand_date(),
            "approver": rand_approver(),
            "category": random.choice(CATEGORIES),
            "desc": f"{asm_name} bileşen grubu",
            "notes": "",
        })
        idx += 1

        for bi in range(10):
            prefix_b = random.choice(ASSEMBLY_PREFIXES)
            sub_name = f"{prefix_b} Alt Montaj {si+1}.{ai+1}.{bi+1}"
            pn2 = rand_pn("SUB", idx)
            rows.append({
                "level": 2,
                "name": sub_name,
                "pn": pn2,
                "sn": rand_sn(idx),
                "qty": random.choice([1, 1, 2, 3]),
                "jira": rand_jira(),
                "weight": round(random.uniform(0.01, 2.0), 4),
                "material": random.choice(MATERIALS),
                "supplier": random.choice(SUPPLIERS),
                "rev": f"Rev {random.randint(1,3):02d}",
                "status": random.choice(STATUSES),
                "criticality": random.choice(CRITICALITIES),
                "mtbf": random.randint(5000, 200000),
                "unit_price": round(random.uniform(10, 2000), 2),
                "currency": random.choice(CURRENCIES),
                "last_rev": rand_date(),
                "approver": rand_approver(),
                "category": random.choice(CATEGORIES),
                "desc": f"Alt montaj parça grubu",
                "notes": "",
            })
            idx += 1

            # Level 3: 9 parça
            for ci in range(9):
                part_name = f"{random.choice(PART_TYPES)} {si+1}.{ai+1}.{bi+1}.{ci+1}"
                pn3 = rand_pn("PRT", idx)
                rows.append({
                    "level": 3,
                    "name": part_name,
                    "pn": pn3,
                    "sn": rand_sn(idx),
                    "qty": random.choice([1, 1, 1, 2, 4, 8, 16]),
                    "jira": rand_jira() if random.random() > 0.5 else "",
                    "weight": round(random.uniform(0.001, 0.5), 5),
                    "material": random.choice(MATERIALS),
                    "supplier": random.choice(SUPPLIERS),
                    "rev": f"Rev {random.randint(1,2):02d}",
                    "status": random.choice(STATUSES),
                    "criticality": random.choice(CRITICALITIES),
                    "mtbf": random.randint(10000, 1000000),
                    "unit_price": round(random.uniform(0.5, 500), 2),
                    "currency": random.choice(CURRENCIES),
                    "last_rev": rand_date(),
                    "approver": rand_approver(),
                    "category": random.choice(CATEGORIES),
                    "desc": f"Standart/özel parça",
                    "notes": f"Lot: {random.randint(1000,9999)}" if random.random() > 0.7 else "",
                })
                idx += 1

print(f"Toplam satır: {len(rows)}")

# ─── Excel Yaz ──────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "BOM"

HEADERS = [
    "Level", "Parça Adı", "Parça No", "Seri No", "Adet", "Jira Key",
    "Ağırlık(kg)", "Malzeme", "Tedarikçi", "Revizyon", "Durum",
    "Kritiklik", "MTBF(saat)", "Birim Fiyat", "Para Birimi",
    "Son Revizyon", "Onaylayan", "Kategori", "Açıklama", "Notlar"
]

# Header stili
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="1E1E2E")
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, bottom=thin)

ws.append(HEADERS)
for col_idx, col_name in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Satır renkleri: derinliğe göre
DEPTH_FILLS = {
    0: PatternFill("solid", fgColor="2D1B69"),  # koyu mor - subsystem
    1: PatternFill("solid", fgColor="1A3A4A"),  # koyu mavi - assembly
    2: PatternFill("solid", fgColor="1A2A1A"),  # koyu yeşil - sub-assembly
    3: None,                                      # varsayılan - part
}
DEPTH_FONTS = {
    0: Font(bold=True, color="C0ADFF", size=10),
    1: Font(bold=True, color="7FC8E8", size=10),
    2: Font(color="7FCC7F", size=9),
    3: Font(color="CCCCCC", size=9),
}

for row_data in rows:
    depth = row_data["level"]
    ws.append([
        row_data["level"], row_data["name"], row_data["pn"], row_data["sn"],
        row_data["qty"], row_data["jira"], row_data["weight"], row_data["material"],
        row_data["supplier"], row_data["rev"], row_data["status"],
        row_data["criticality"], row_data["mtbf"], row_data["unit_price"],
        row_data["currency"], row_data["last_rev"], row_data["approver"],
        row_data["category"], row_data["desc"], row_data["notes"]
    ])
    row_num = ws.max_row
    fill = DEPTH_FILLS.get(depth)
    font = DEPTH_FONTS.get(depth, Font(color="CCCCCC", size=9))
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        if fill:
            cell.fill = fill
        cell.font = font
        if depth == 0:
            # Level indent hint via first column alignment
            cell.alignment = Alignment(horizontal="left")

# Sütun genişlikleri
COL_WIDTHS = [7, 45, 20, 16, 7, 12, 12, 14, 14, 10, 12, 10, 12, 12, 8, 14, 14, 14, 35, 20]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

ws.freeze_panes = "B2"
ws.row_dimensions[1].height = 22

output_path = r"D:\projeler\bom-manager\test_bom_20k.xlsx"
wb.save(output_path)
print(f"Kaydedildi: {output_path}")
